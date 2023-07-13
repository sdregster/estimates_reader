from openpyxl import load_workbook

from etc.entities import Chapter, Material, MiM, Subchapter, Work


class EstimateBaseTemplate:
    CHAPTERS = []
    FOOTER = None

    def __init__(self, file_path: str) -> None:
        self.__wb = load_workbook(file_path, data_only=True)
        self.__ws = self.__wb.active
        self.estimate_file_name = file_path.split("\\")[-1]
        self._get_estimate_keystone_data()
        self.rows = []

        self.chapter_start = None
        self.footer_start = None

    def _get_estimate_keystone_data(self):
        def _get_group_code(value: str) -> str:
            temp_result = value[value.rfind("(") + 1 :]
            return temp_result[: temp_result.rfind(")")]

        self.estimate_total_number = (
            str(self.__ws.cell(9, 6).value).split("№")[1].strip()
        )
        if " изм." in self.estimate_total_number:
            (
                self.estimate_number,
                self.estimate_version,
            ) = self.estimate_total_number.split(" изм.")
        else:
            self.estimate_number, self.estimate_version = self.estimate_total_number, 0
        if self.__ws.cell(12, 4).value:
            self.estimate_work_name = (
                str(self.__ws.cell(12, 4).value).replace("\n", "").strip()
            )
        else:
            self.estimate_work_name = (
                str(self.__ws.cell(12, 3).value).replace("\n", "").strip()
            )

        self.estimate_reason = (
            str(self.__ws.cell(15, 3).value).split("Основание: ")[1].strip()
        )
        self.estimate_group_code = self.estimate_reason
        self.estimate_cipher = self.estimate_reason
        # self.estimate_time_period = (
        #     str(self.__ws.cell(19, 6).value).replace("_", "").strip()
        # ) #TODO

        for row in range(1, self.__ws.max_row + 1):
            target_value = self.__ws.cell(row, 1).value
            time_perion_search_value = self.__ws.cell(row, 3).value
            if isinstance(target_value, str) and "№ пп" in target_value:
                self.content_start_from = row + 4
                # elif isinstance(target_value, str) and "ФОТ" in target_value:
                self.estimate_wage_fund = self.__ws.cell(row, 8).value
            elif isinstance(target_value, str) and "ВСЕГО по смете" in target_value:
                self.estimate_cost = self.__ws.cell(row, 8).value
                self.estimate_wage_fund = self.__ws.cell(row - 3, 8).value
                self.estimate_laboriousness = self.__ws.cell(row, 13).value
            elif (
                isinstance(time_perion_search_value, str)
                and "Составлен(а) в текущих (прогнозных) ценах по состоянию на"
                in time_perion_search_value
            ):
                self.estimate_time_period = str(self.__ws.cell(row, 3).value).split(
                    "ценах по состоянию на"
                )[1]

            # Поиск разделов
            elif (
                isinstance(target_value, str)
                and "Итого прямые затраты по разделу в текущих ценах" in target_value
            ):
                self.chapter_start = row
            elif (
                isinstance(target_value, str)
                and "Итого по разделу" in target_value
                and self.chapter_start
            ):
                EstimateBaseTemplate.CHAPTERS.append((self.chapter_start, row))
                self.chapter_start = None

        # Поиск подвала
        if EstimateBaseTemplate.CHAPTERS:
            for row in range(
                EstimateBaseTemplate.CHAPTERS[-1][1], self.__ws.max_row + 1
            ):
                target_value = self.__ws.cell(row, 1).value
                if isinstance(target_value, str) and "ИТОГИ ПО СМЕТЕ:" in target_value:
                    self.footer_start = row
                elif (
                    isinstance(target_value, str)
                    and "ВСЕГО по смете" in target_value
                    and self.footer_start
                ):
                    EstimateBaseTemplate.FOOTER = (self.footer_start, row)
                    self.footer_start = None
        else:
            for row in range(1, self.__ws.max_row + 1):
                target_value = self.__ws.cell(row, 1).value
                if (
                    isinstance(target_value, str)
                    and "Итого прямые затраты по смете в текущих ценах" in target_value
                ):
                    self.footer_start = row
                elif (
                    isinstance(target_value, str)
                    and "ВСЕГО по смете" in target_value
                    and self.footer_start
                ):
                    EstimateBaseTemplate.FOOTER = (self.footer_start, row)
                    self.footer_start = None

    def read_rows(self):
        rows_black_list = []
        if EstimateBaseTemplate.CHAPTERS:
            for chapter in EstimateBaseTemplate.CHAPTERS:
                rows_black_list.extend(range(chapter[0], chapter[1] + 1))
        if EstimateBaseTemplate.FOOTER:
            rows_black_list.extend(
                range(
                    EstimateBaseTemplate.FOOTER[0], EstimateBaseTemplate.FOOTER[1] + 1
                )
            )

        for row in range(self.content_start_from, self.__ws.max_row + 1):
            if row not in rows_black_list and row <= rows_black_list[-1]:
                current_row_values = [
                    self.__ws.cell(row, col).value
                    for col in range(1, self.__ws.max_column + 1)
                ]

                nn_value = self.__ws.cell(row, 1).value
                reason_value = self.__ws.cell(row, 2).value
                unit_value = self.__ws.cell(row, 4).value

                # Раздел
                if isinstance(nn_value, str) and "Раздел" in nn_value:
                    self.rows.append(Chapter(current_row_values))

                # Подраздел
                elif (
                    isinstance(nn_value, str)
                    and "Раздел" not in nn_value
                    and not unit_value
                ):
                    self.has_subchapters = True
                    self.rows.append(Subchapter(current_row_values))

                # Работа
                elif (
                    isinstance(reason_value, str)
                    and nn_value
                    and "ГЭСН" in reason_value
                ):
                    self.rows.append(Work(current_row_values))

                # Материал
                elif (
                    isinstance(reason_value, str)
                    and nn_value
                    and nn_value != "Н"
                    and "ГЭСН" not in reason_value
                ):
                    self.rows.append(Material(current_row_values))

                # МиМ
                elif (
                    isinstance(reason_value, str)
                    and not nn_value
                    and "маш.час" in unit_value
                ):
                    self.rows.append(MiM(current_row_values))
