from modules.estimate_data_collector import EstimateBaseTemplate
from etc.entities import Chapter, Subchapter, Work, Material, MiM
from openpyxl import load_workbook
from datetime import date

from pprint import pprint


class NeosintezTemplate:
    def __init__(self) -> None:
        self.__wb = load_workbook("data/template.xlsx")
        self.__ws = self.__wb.active
        self.cursor = 2

    def _write_header(self, obj: EstimateBaseTemplate, class_levels: dict):
        class_name = "Смета"
        self.__ws.cell(self.cursor, 1).value = obj.estimate_total_number
        self.__ws.cell(self.cursor, 2).value = class_levels[class_name]
        self.__ws.cell(self.cursor, 4).value = class_name
        self.__ws.cell(self.cursor, 9).value = obj.estimate_total_number
        self.__ws.cell(self.cursor, 23).value = obj.estimate_work_name
        self.__ws.cell(self.cursor, 24).value = obj.estimate_number
        self.__ws.cell(self.cursor, 25).value = int(obj.estimate_version)
        self.__ws.cell(self.cursor, 26).value = obj.estimate_reason
        self.__ws.cell(self.cursor, 27).value = obj.estimate_cipher
        self.__ws.cell(self.cursor, 28).value = obj.estimate_cost
        self.__ws.cell(self.cursor, 29).value = obj.estimate_wage_fund
        self.__ws.cell(self.cursor, 30).value = obj.estimate_laboriousness
        self.__ws.cell(self.cursor, 31).value = obj.estimate_time_period
        self.__ws.cell(self.cursor, 32).value = obj.estimate_file_name

        self.cursor += 1

    def _write_chapter(self, estimate: EstimateBaseTemplate, class_levels: dict, obj: Chapter):
        class_name = "Раздел сметы"

        self.__ws.cell(self.cursor, 1).value = estimate.estimate_total_number
        self.__ws.cell(self.cursor, 2).value = class_levels[class_name]
        self.__ws.cell(self.cursor, 4).value = class_name
        self.__ws.cell(self.cursor, 5).value = obj.name
        self.__ws.cell(self.cursor, 9).value = obj.name

        self.cursor += 1

    def _write_subchapter(self, estimate: EstimateBaseTemplate, class_levels: dict, obj: Subchapter):
        class_name = "Подраздел сметы"

        self.__ws.cell(self.cursor, 1).value = estimate.estimate_total_number
        self.__ws.cell(self.cursor, 2).value = class_levels[class_name]
        self.__ws.cell(self.cursor, 4).value = class_name
        self.__ws.cell(self.cursor, 5).value = obj.name
        self.__ws.cell(self.cursor, 9).value = obj.name

        self.cursor += 1

    def _write_work(self, estimate: EstimateBaseTemplate, class_levels: dict, obj: Work):
        class_name = "Строка сметы"

        self.__ws.cell(self.cursor, 1).value = estimate.estimate_total_number
        self.__ws.cell(self.cursor, 2).value = class_levels[class_name]
        self.__ws.cell(self.cursor, 4).value = class_name
        self.__ws.cell(self.cursor, 8).value = "Работа"
        self.__ws.cell(self.cursor, 9).value = obj.total_name
        self.__ws.cell(self.cursor, 10).value = obj.index
        self.__ws.cell(self.cursor, 11).value = obj.reason
        self.__ws.cell(self.cursor, 12).value = obj.name
        self.__ws.cell(self.cursor, 13).value = obj.unit
        self.__ws.cell(self.cursor, 14).value = obj.amount
        self.__ws.cell(self.cursor, 15).value = obj.cost_per_unit
        self.__ws.cell(self.cursor, 16).value = obj.total_cost
        self.__ws.cell(self.cursor, 17).value = obj.total_wage
        self.__ws.cell(self.cursor, 18).value = obj.mim_cost
        self.__ws.cell(self.cursor, 19).value = obj.mim_wage
        self.__ws.cell(self.cursor, 20).value = obj.materials_cost
        self.__ws.cell(self.cursor, 21).value = obj.laboriousness
        self.__ws.cell(self.cursor, 22).value = obj.mim_laboriousness

        self.cursor += 1

    def _write_material(self, estimate: EstimateBaseTemplate, class_levels: dict, obj: Material):
        class_name = "Строка сметы"

        self.__ws.cell(self.cursor, 1).value = estimate.estimate_total_number
        self.__ws.cell(self.cursor, 2).value = class_levels[class_name]
        self.__ws.cell(self.cursor, 4).value = class_name
        self.__ws.cell(self.cursor, 8).value = "МТР-Материалы"
        self.__ws.cell(self.cursor, 9).value = obj.total_name
        self.__ws.cell(self.cursor, 10).value = obj.index
        self.__ws.cell(self.cursor, 11).value = obj.reason
        self.__ws.cell(self.cursor, 12).value = obj.name
        self.__ws.cell(self.cursor, 13).value = obj.unit
        self.__ws.cell(self.cursor, 14).value = obj.amount
        self.__ws.cell(self.cursor, 15).value = obj.cost_per_unit
        self.__ws.cell(self.cursor, 16).value = obj.total_cost
        self.__ws.cell(self.cursor, 17).value = obj.total_wage
        self.__ws.cell(self.cursor, 18).value = obj.mim_cost
        self.__ws.cell(self.cursor, 19).value = obj.mim_wage
        self.__ws.cell(self.cursor, 20).value = obj.materials_cost
        self.__ws.cell(self.cursor, 21).value = obj.laboriousness
        self.__ws.cell(self.cursor, 22).value = obj.mim_laboriousness

        self.cursor += 1

    def _write_mim(self, estimate: EstimateBaseTemplate, class_levels: dict, obj: MiM):
        class_name = "МиМ сметы"

        self.__ws.cell(self.cursor, 1).value = estimate.estimate_total_number
        self.__ws.cell(self.cursor, 2).value = class_levels[class_name]
        self.__ws.cell(self.cursor, 4).value = class_name
        self.__ws.cell(self.cursor, 8).value = "МиМ"
        self.__ws.cell(self.cursor, 9).value = obj.name
        self.__ws.cell(self.cursor, 11).value = obj.reason
        self.__ws.cell(self.cursor, 12).value = obj.name
        self.__ws.cell(self.cursor, 13).value = obj.unit
        self.__ws.cell(self.cursor, 14).value = obj.amount
        self.__ws.cell(self.cursor, 15).value = obj.cost_per_unit
        self.__ws.cell(self.cursor, 16).value = obj.total_cost
        self.__ws.cell(self.cursor, 17).value = obj.total_wage
        self.__ws.cell(self.cursor, 18).value = obj.mim_cost
        self.__ws.cell(self.cursor, 19).value = obj.mim_wage
        self.__ws.cell(self.cursor, 20).value = obj.materials_cost
        self.__ws.cell(self.cursor, 21).value = obj.laboriousness
        self.__ws.cell(self.cursor, 22).value = obj.mim_laboriousness

        self.cursor += 1

    def _finish(self):
        self.__wb.save(f"output/{date.today()}_nt_template.xlsx")
        self.__wb.close()

    def export(self, estimates_list: list[EstimateBaseTemplate]):
        for estimate in estimates_list:
            class_levels = (
                {
                    "Смета": 1,
                    "Раздел сметы": 2,
                    "Подраздел сметы": 3,
                    "Строка сметы": 4,
                    "МиМ сметы": 5,
                }
                if estimate.has_subchapters
                else {
                    "Смета": 1,
                    "Раздел сметы": 2,
                    "Строка сметы": 3,
                    "МиМ сметы": 4,
                }
            )

            self._write_header(estimate, class_levels)

            for row in estimate.rows:
                if isinstance(row, Chapter):
                    self._write_chapter(estimate, class_levels, row)

                if isinstance(row, Subchapter):
                    self._write_subchapter(estimate, class_levels, row)

                if isinstance(row, Work):
                    self._write_work(estimate, class_levels, row)

                if isinstance(row, Material):
                    self._write_material(estimate, class_levels, row)

                if isinstance(row, MiM):
                    self._write_mim(estimate, class_levels, row)

            # finish
            self._finish()
